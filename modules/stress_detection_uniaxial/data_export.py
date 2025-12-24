"""
应力场测绘模块 - 数据验证和导出
负责数据验证、CSV/Excel/HDF5导出
"""

import os
import csv
import json
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional

from .field_database import FieldDatabaseManager
from .field_hdf5 import FieldExperimentHDF5


class DataValidator:
    """数据验证器类"""
    
    # 验证阈值
    STRESS_MIN = -1000  # MPa
    STRESS_MAX = 1000   # MPa
    TIME_DIFF_MIN = -1000  # ns
    TIME_DIFF_MAX = 1000   # ns
    NEIGHBOR_DIFF_MAX = 200  # MPa（与相邻点差异阈值）
    MIN_SHAPE_AREA = 100   # mm²
    MIN_POINT_COUNT = 3
    MIN_R_SQUARED = 0.95
    
    @staticmethod
    def validate_point_data(point: Dict[str, Any], 
                           neighbors: List[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        验证单个测点数据
        
        Args:
            point: 测点数据
            neighbors: 相邻测点列表 (可选)
        
        Returns:
            dict: {"is_valid": bool, "is_suspicious": bool, "warnings": [...]}
        """
        warnings = []
        is_suspicious = False
        
        stress = point.get('stress_value')
        time_diff = point.get('time_diff')
        
        # 验证应力值范围
        if stress is not None:
            if not (DataValidator.STRESS_MIN <= stress <= DataValidator.STRESS_MAX):
                is_suspicious = True
                warnings.append(f"应力值 {stress:.1f} MPa 超出正常范围 [{DataValidator.STRESS_MIN}, {DataValidator.STRESS_MAX}]")
        
        # 验证时间差范围
        if time_diff is not None:
            if not (DataValidator.TIME_DIFF_MIN <= time_diff <= DataValidator.TIME_DIFF_MAX):
                is_suspicious = True
                warnings.append(f"时间差 {time_diff:.2f} ns 超出正常范围")
        
        # 检查与相邻点的连续性
        if neighbors and stress is not None:
            for neighbor in neighbors:
                neighbor_stress = neighbor.get('stress_value')
                if neighbor_stress is not None:
                    diff = abs(stress - neighbor_stress)
                    if diff > DataValidator.NEIGHBOR_DIFF_MAX:
                        is_suspicious = True
                        warnings.append(f"与相邻点应力差异过大 ({diff:.1f} MPa)")
                        break
        
        return {
            "is_valid": len(warnings) == 0,
            "is_suspicious": is_suspicious,
            "warnings": warnings
        }
    
    @staticmethod
    def validate_experiment_config(config: Dict[str, Any]) -> Dict[str, Any]:
        """
        验证实验配置
        
        Args:
            config: 实验配置
        
        Returns:
            dict: {"is_valid": bool, "warnings": [...], "errors": [...]}
        """
        warnings = []
        errors = []
        
        # 验证形状面积
        shape_config = config.get('shape_config', {})
        if shape_config:
            from .shape_utils import ShapeUtils
            validation = ShapeUtils.validate_shape(shape_config)
            if not validation['is_valid']:
                errors.append(validation.get('error', '形状配置无效'))
            elif validation['area'] < DataValidator.MIN_SHAPE_AREA:
                warnings.append(f"形状面积 {validation['area']:.1f} mm² 较小")
        
        # 验证测点数量
        point_count = config.get('point_count', 0)
        if point_count < DataValidator.MIN_POINT_COUNT:
            warnings.append(f"测点数量 {point_count} 较少，建议 ≥ {DataValidator.MIN_POINT_COUNT}")
        
        # 验证标定数据
        calibration = config.get('calibration', {})
        r_squared = calibration.get('r_squared')
        if r_squared is not None and r_squared < DataValidator.MIN_R_SQUARED:
            warnings.append(f"标定拟合优度 R²={r_squared:.4f} 较低")
        
        return {
            "is_valid": len(errors) == 0,
            "warnings": warnings,
            "errors": errors
        }


class DataExporter:
    """数据导出器类"""
    
    def __init__(self, db: FieldDatabaseManager):
        """
        初始化导出器
        
        Args:
            db: 数据库管理器
        """
        self.db = db
    
    def _merge_hdf5_config(self, exp_id: str, exp_data: Dict) -> Dict:
        """
        从HDF5读取config_snapshot并合并到exp_data
        
        Args:
            exp_id: 实验ID
            exp_data: 实验数据字典
        
        Returns:
            合并后的exp_data
        """
        try:
            hdf5 = FieldExperimentHDF5(exp_id)
            if hdf5.file_exists():
                hdf5_config = hdf5.load_config_snapshot()
                if hdf5_config.get('success') and hdf5_config.get('data'):
                    # 合并HDF5中的配置到exp_data
                    if not exp_data.get('config_snapshot'):
                        exp_data['config_snapshot'] = {}
                    exp_data['config_snapshot'].update(hdf5_config['data'])
        except Exception as e:
            # 如果HDF5读取失败，继续使用数据库中的数据
            print(f"Warning: Failed to load HDF5 config: {e}")
        return exp_data
    
    def _infer_grid_params_from_points(self, points: List[Dict], shape_config: Dict) -> Dict:
        """
        从测点数据反推网格布点参数（用于历史数据兼容）
        
        Args:
            points: 测点列表
            shape_config: 形状配置
        
        Returns:
            推断出的布点参数
        """
        if not points:
            return {}
        
        try:
            # 获取所有 X 和 Y 坐标
            x_coords = sorted(set(p.get('x_coord', 0) for p in points if p.get('x_coord') is not None))
            y_coords = sorted(set(p.get('y_coord', 0) for p in points if p.get('y_coord') is not None))
            
            if not x_coords or not y_coords:
                return {}
            
            params = {}
            
            # 计算行列数
            params['rows'] = len(y_coords)
            params['cols'] = len(x_coords)
            
            # 计算边距（从形状边界到最外侧测点的距离）
            if shape_config.get('type') == 'rectangle':
                width = shape_config.get('width', 0)
                height = shape_config.get('height', 0)
                
                if width > 0 and height > 0:
                    params['margin_left'] = round(min(x_coords), 2)
                    params['margin_right'] = round(width - max(x_coords), 2)
                    params['margin_top'] = round(min(y_coords), 2)
                    params['margin_bottom'] = round(height - max(y_coords), 2)
            
            return params
        except Exception as e:
            print(f"Warning: Failed to infer grid params: {e}")
            return {}
    
    def export_to_csv(self, exp_id: str, output_path: str = None,
                     include_quality: bool = True) -> Dict[str, Any]:
        """
        导出为CSV文件
        
        Args:
            exp_id: 实验ID
            output_path: 输出路径 (可选)
            include_quality: 是否包含质量评分
        
        Returns:
            dict: {"success": bool, "file_path": str}
        """
        try:
            # 加载实验数据
            result = self.db.load_experiment(exp_id)
            if not result['success']:
                return result
            
            exp_data = result['data']['experiment']
            points = result['data']['points']
            
            # 尝试从HDF5读取完整的config_snapshot（包含layout配置）
            exp_data = self._merge_hdf5_config(exp_id, exp_data)
            
            # 生成输出路径
            if output_path is None:
                output_dir = 'data/uniaxial_field/exports'
                os.makedirs(output_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(output_dir, f'{exp_id}_data_{timestamp}.csv')
            
            # 确定坐标类型
            shape_config = exp_data.get('shape_config', {})
            use_polar = (shape_config.get('type') == 'circle' and 
                        any(p.get('r_coord') is not None for p in points))
            
            # 计算统计信息
            measured_points = [p for p in points if p.get('status') == 'measured']
            stress_values = [p['stress_value'] for p in measured_points if p.get('stress_value') is not None]
            
            # 写入CSV
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # 写入元数据
                writer.writerow(['实验ID', exp_id])
                writer.writerow(['实验名称', exp_data.get('name', '')])
                writer.writerow(['状态', exp_data.get('status', '')])
                writer.writerow(['创建时间', exp_data.get('created_at', '')])
                writer.writerow(['完成时间', exp_data.get('completed_at', '')])
                writer.writerow(['试件材料', exp_data.get('sample_material', '')])
                writer.writerow(['试件厚度(mm)', exp_data.get('sample_thickness', '')])
                writer.writerow(['楔块角度(°)', exp_data.get('wedge_angle', '')])
                writer.writerow(['应力方向', exp_data.get('stress_direction', '')])
                writer.writerow(['实验目的', exp_data.get('test_purpose', '')])
                writer.writerow(['操作员', exp_data.get('operator', '')])
                writer.writerow(['环境温度(°C)', exp_data.get('temperature', '')])
                writer.writerow(['环境湿度(%)', exp_data.get('humidity', '')])
                writer.writerow(['示波器型号', exp_data.get('scope_model', '')])
                writer.writerow(['探头型号', exp_data.get('probe_model', '')])
                writer.writerow(['备注', exp_data.get('notes', '')])
                writer.writerow(['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([])
                
                # 写入标定信息
                writer.writerow(['--- 标定信息 ---'])
                writer.writerow(['标定实验ID', exp_data.get('calibration_exp_id', '')])
                writer.writerow(['标定方向', exp_data.get('calibration_direction', '')])
                writer.writerow(['标定系数(MPa/ns)', exp_data.get('calibration_k', '')])
                
                # 从HDF5配置快照中获取更详细的标定信息
                calibration_config = exp_data.get('config_snapshot', {}).get('calibration', {})
                if calibration_config:
                    writer.writerow(['拟合优度(R²)', calibration_config.get('r_squared', '')])
                    writer.writerow(['斜率', calibration_config.get('slope', '')])
                    writer.writerow(['截距', calibration_config.get('intercept', '')])
                
                writer.writerow([])
                
                # 写入基准点信息
                baseline_point_id = exp_data.get('baseline_point_id')
                baseline_stress = exp_data.get('baseline_stress')
                if baseline_point_id is not None or baseline_stress is not None:
                    writer.writerow(['--- 基准点信息 ---'])
                    writer.writerow(['基准点ID', baseline_point_id if baseline_point_id else ''])
                    writer.writerow(['基准应力(MPa)', baseline_stress if baseline_stress else ''])
                    writer.writerow([])
                
                # 写入形状配置
                writer.writerow(['--- 形状配置 ---'])
                shape_type_map = {'rectangle': '矩形', 'circle': '圆形', 'polygon': '多边形'}
                shape_type = shape_config.get('type', '')
                writer.writerow(['形状类型', shape_type_map.get(shape_type, shape_type)])
                
                if shape_type == 'rectangle':
                    writer.writerow(['宽度(mm)', shape_config.get('width', '')])
                    writer.writerow(['高度(mm)', shape_config.get('height', '')])
                elif shape_type == 'circle':
                    writer.writerow(['半径(mm)', shape_config.get('radius', '')])
                    center = shape_config.get('center', {})
                    writer.writerow(['圆心X(mm)', center.get('x', '')])
                    writer.writerow(['圆心Y(mm)', center.get('y', '')])
                elif shape_type == 'polygon':
                    vertices = shape_config.get('vertices', [])
                    writer.writerow(['顶点数', len(vertices)])
                
                # 写入孔洞信息
                modifiers = shape_config.get('modifiers', [])
                holes = [m for m in modifiers if m.get('op') == 'subtract']
                if holes:
                    writer.writerow(['孔洞数量', len(holes)])
                    for i, hole in enumerate(holes, 1):
                        hole_shape = hole.get('shape', 'circle')
                        if hole_shape == 'circle':
                            writer.writerow([f'孔洞{i}类型', '圆形'])
                            writer.writerow([f'孔洞{i}圆心X(mm)', hole.get('centerX', '')])
                            writer.writerow([f'孔洞{i}圆心Y(mm)', hole.get('centerY', '')])
                            writer.writerow([f'孔洞{i}半径(mm)', hole.get('radius', '')])
                        elif hole_shape == 'rectangle':
                            writer.writerow([f'孔洞{i}类型', '矩形'])
                            writer.writerow([f'孔洞{i}中心X(mm)', hole.get('centerX', '')])
                            writer.writerow([f'孔洞{i}中心Y(mm)', hole.get('centerY', '')])
                            writer.writerow([f'孔洞{i}宽度(mm)', hole.get('width', '')])
                            writer.writerow([f'孔洞{i}高度(mm)', hole.get('height', '')])
                
                writer.writerow([])
                
                # 写入布点配置
                writer.writerow(['--- 布点配置 ---'])
                layout_config = exp_data.get('config_snapshot', {}).get('layout', {})
                layout_type_map = {'grid': '网格布点', 'polar': '极坐标布点', 'custom': '自定义布点'}
                layout_type = layout_config.get('type', '')
                writer.writerow(['布点方式', layout_type_map.get(layout_type, layout_type)])
                
                # 获取布点参数（可能在params子字典中）
                layout_params = layout_config.get('params', {})
                
                # 如果 params 为空，尝试从测点数据反推
                if not layout_params and points and layout_type == 'grid':
                    layout_params = self._infer_grid_params_from_points(points, shape_config)
                
                if layout_type == 'grid':
                    # 兼容性适配：前端可能使用不同的参数名
                    # 前端: rows, cols, margin_left, margin_right, margin_top, margin_bottom
                    # 后端: spacing_x, spacing_y, margins: {left, right, top, bottom}
                    rows = layout_params.get('rows', '')
                    cols = layout_params.get('cols', '')
                    spacing_x = layout_params.get('spacing_x') or layout_params.get('cols', '')
                    spacing_y = layout_params.get('spacing_y') or layout_params.get('rows', '')
                    
                    # 如果有行列数，优先显示
                    if rows:
                        writer.writerow(['行数', rows])
                    if cols:
                        writer.writerow(['列数', cols])
                    
                    # 边距信息
                    margins = layout_params.get('margins', {})
                    margin_left = margins.get('left') if margins else layout_params.get('margin_left', '')
                    margin_right = margins.get('right') if margins else layout_params.get('margin_right', '')
                    margin_top = margins.get('top') if margins else layout_params.get('margin_top', '')
                    margin_bottom = margins.get('bottom') if margins else layout_params.get('margin_bottom', '')
                    
                    if margin_left or margin_right or margin_top or margin_bottom:
                        writer.writerow(['左边距(mm)', margin_left])
                        writer.writerow(['右边距(mm)', margin_right])
                        writer.writerow(['上边距(mm)', margin_top])
                        writer.writerow(['下边距(mm)', margin_bottom])
                elif layout_type == 'polar':
                    writer.writerow(['半径步长(mm)', layout_params.get('radius_step', '')])
                    writer.writerow(['角度步长(°)', layout_params.get('angle_step', '')])
                    writer.writerow(['起始半径(mm)', layout_params.get('start_radius', '')])
                writer.writerow([])
                
                # 写入统计信息
                writer.writerow(['--- 统计信息 ---'])
                writer.writerow(['总测点数', len(points)])
                writer.writerow(['已测量', len(measured_points)])
                writer.writerow(['待测量', len([p for p in points if p.get('status') == 'pending'])])
                writer.writerow(['已跳过', len([p for p in points if p.get('status') == 'skipped'])])
                
                if stress_values:
                    writer.writerow([])
                    writer.writerow(['应力最小值(MPa)', f'{min(stress_values):.2f}'])
                    writer.writerow(['应力最大值(MPa)', f'{max(stress_values):.2f}'])
                    writer.writerow(['应力平均值(MPa)', f'{np.mean(stress_values):.2f}'])
                    writer.writerow(['应力标准差(MPa)', f'{np.std(stress_values):.2f}'])
                    writer.writerow(['应力范围(MPa)', f'{max(stress_values) - min(stress_values):.2f}'])
                
                writer.writerow([])
                
                # 写入云图配置（如果有）
                metadata = result['data'].get('metadata')
                if metadata:
                    writer.writerow(['--- 云图配置 ---'])
                    writer.writerow(['插值方法', metadata.get('interpolation_method', '')])
                    writer.writerow(['网格分辨率', metadata.get('grid_resolution', '')])
                    writer.writerow(['色标范围最小值(MPa)', metadata.get('vmin', '')])
                    writer.writerow(['色标范围最大值(MPa)', metadata.get('vmax', '')])
                    writer.writerow([])
                
                writer.writerow(['--- 测点数据 ---'])
                
                # 写入表头
                if use_polar:
                    headers = ['#', 'R(mm)', 'θ(°)', 'Δt(ns)', 'σ(MPa)', '状态']
                else:
                    headers = ['#', 'X(mm)', 'Y(mm)', 'Δt(ns)', 'σ(MPa)', '状态']
                
                if include_quality:
                    headers.extend(['质量评分', 'SNR(dB)', '可疑'])
                
                # 添加跳过原因列
                headers.append('跳过原因')
                
                writer.writerow(headers)
                
                # 写入数据
                for point in points:
                    if use_polar:
                        row = [
                            point['point_index'],
                            f"{point.get('r_coord', 0):.2f}",
                            f"{point.get('theta_coord', 0):.1f}",
                            f"{point.get('time_diff', 0):.3f}" if point.get('time_diff') else '',
                            f"{point.get('stress_value', 0):.2f}" if point.get('stress_value') else '',
                            point.get('status', '')
                        ]
                    else:
                        row = [
                            point['point_index'],
                            f"{point.get('x_coord', 0):.2f}",
                            f"{point.get('y_coord', 0):.2f}",
                            f"{point.get('time_diff', 0):.3f}" if point.get('time_diff') else '',
                            f"{point.get('stress_value', 0):.2f}" if point.get('stress_value') else '',
                            point.get('status', '')
                        ]
                    
                    if include_quality:
                        row.extend([
                            f"{point.get('quality_score', 0):.2f}" if point.get('quality_score') else '',
                            f"{point.get('snr', 0):.1f}" if point.get('snr') else '',
                            '是' if point.get('is_suspicious') else ''
                        ])
                    
                    # 添加跳过原因
                    row.append(point.get('skip_reason', ''))
                    
                    writer.writerow(row)
            
            return {
                "success": True,
                "error_code": 0,
                "file_path": output_path,
                "message": f"已导出 {len(points)} 个测点数据"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error_code": 5001,
                "message": f"CSV导出失败: {str(e)}"
            }
    
    def export_to_excel(self, exp_id: str, output_path: str = None, 
                       single_sheet: bool = False) -> Dict[str, Any]:
        """
        导出为Excel文件
        
        Args:
            exp_id: 实验ID
            output_path: 输出路径 (可选)
            single_sheet: 是否使用单表格式 (默认False，使用多表格式)
        
        Returns:
            dict: {"success": bool, "file_path": str}
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return {
                "success": False,
                "error_code": 5010,
                "message": "openpyxl库未安装，无法导出Excel"
            }
        
        try:
            # 加载实验数据
            result = self.db.load_experiment(exp_id)
            if not result['success']:
                return result
            
            exp_data = result['data']['experiment']
            points = result['data']['points']
            
            # 尝试从HDF5读取完整的config_snapshot（包含layout配置）
            exp_data = self._merge_hdf5_config(exp_id, exp_data)
            
            # 生成输出路径
            if output_path is None:
                output_dir = 'data/uniaxial_field/exports'
                os.makedirs(output_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                format_suffix = 'single' if single_sheet else 'multi'
                output_path = os.path.join(output_dir, f'{exp_id}_data_{format_suffix}_{timestamp}.xlsx')
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            if single_sheet:
                # 单表格式：所有数据在一个表中
                ws = wb.active
                ws.title = '实验数据'
                
                current_row = 1
                
                # 写入实验信息（作为表头）
                meta_data = [
                    ('实验ID', exp_id),
                    ('实验名称', exp_data.get('name', '')),
                    ('创建时间', exp_data.get('created_at', '')),
                    ('试件材料', exp_data.get('sample_material', '')),
                    ('试件厚度(mm)', exp_data.get('sample_thickness', '')),
                    ('楔块角度(°)', exp_data.get('wedge_angle', '')),
                    ('操作员', exp_data.get('operator', '')),
                ]
                
                for key, value in meta_data:
                    ws.cell(row=current_row, column=1, value=key).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=value)
                    current_row += 1
                
                current_row += 1  # 空行
                
                # 写入统计信息
                measured_points = [p for p in points if p.get('status') == 'measured']
                stress_values = [p['stress_value'] for p in measured_points if p.get('stress_value') is not None]
                
                ws.cell(row=current_row, column=1, value='测点统计').font = Font(bold=True)
                current_row += 1
                
                stats_data = [
                    ('总测点数', len(points)),
                    ('已测量', len(measured_points)),
                    ('待测量', len([p for p in points if p.get('status') == 'pending'])),
                ]
                
                if stress_values:
                    stats_data.extend([
                        ('应力最小值(MPa)', round(min(stress_values), 2)),
                        ('应力最大值(MPa)', round(max(stress_values), 2)),
                        ('应力平均值(MPa)', round(np.mean(stress_values), 2)),
                    ])
                
                for key, value in stats_data:
                    ws.cell(row=current_row, column=1, value=key)
                    ws.cell(row=current_row, column=2, value=value)
                    current_row += 1
                
                current_row += 1  # 空行
                
                # 写入测点数据表
                shape_config = exp_data.get('shape_config', {})
                use_polar = shape_config.get('type') == 'circle'
                
                if use_polar:
                    headers = ['#', 'R(mm)', 'θ(°)', 'X(mm)', 'Y(mm)', 'Δt(ns)', 'σ(MPa)', 
                              '状态', '质量评分', 'SNR(dB)']
                else:
                    headers = ['#', 'X(mm)', 'Y(mm)', 'Δt(ns)', 'σ(MPa)', 
                              '状态', '质量评分', 'SNR(dB)']
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                
                current_row += 1
                
                for point in points:
                    if use_polar:
                        data = [
                            point['point_index'],
                            point.get('r_coord'),
                            point.get('theta_coord'),
                            point.get('x_coord'),
                            point.get('y_coord'),
                            point.get('time_diff'),
                            point.get('stress_value'),
                            point.get('status'),
                            point.get('quality_score'),
                            point.get('snr'),
                        ]
                    else:
                        data = [
                            point['point_index'],
                            point.get('x_coord'),
                            point.get('y_coord'),
                            point.get('time_diff'),
                            point.get('stress_value'),
                            point.get('status'),
                            point.get('quality_score'),
                            point.get('snr'),
                        ]
                    
                    for col_idx, value in enumerate(data, 1):
                        ws.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1
                
            else:
                # 多表格式：分为三个工作表
                # Sheet 1: 元数据
                ws_meta = wb.active
                ws_meta.title = '实验信息'
                
                meta_data = [
                    ('实验ID', exp_id),
                    ('实验名称', exp_data.get('name', '')),
                    ('创建时间', exp_data.get('created_at', '')),
                    ('完成时间', exp_data.get('completed_at', '')),
                    ('状态', exp_data.get('status', '')),
                    ('试件材料', exp_data.get('sample_material', '')),
                    ('试件厚度(mm)', exp_data.get('sample_thickness', '')),
                    ('楔块角度(°)', exp_data.get('wedge_angle', '')),
                    ('实验目的', exp_data.get('test_purpose', '')),
                    ('操作员', exp_data.get('operator', '')),
                    ('环境温度(°C)', exp_data.get('temperature', '')),
                    ('环境湿度(%)', exp_data.get('humidity', '')),
                ]
                
                for row_idx, (key, value) in enumerate(meta_data, 1):
                    ws_meta.cell(row=row_idx, column=1, value=key)
                    ws_meta.cell(row=row_idx, column=2, value=value)
                
                # Sheet 2: 测点数据
                ws_points = wb.create_sheet('测点数据')
                
                # 确定坐标类型
                shape_config = exp_data.get('shape_config', {})
                use_polar = shape_config.get('type') == 'circle'
                
                if use_polar:
                    headers = ['#', 'R(mm)', 'θ(°)', 'X(mm)', 'Y(mm)', 'Δt(ns)', 'σ(MPa)', 
                              '状态', '质量评分', 'SNR(dB)', '可疑', '测量时间', '跳过原因']
                else:
                    headers = ['#', 'X(mm)', 'Y(mm)', 'Δt(ns)', 'σ(MPa)', 
                              '状态', '质量评分', 'SNR(dB)', '可疑', '测量时间', '跳过原因']
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_points.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                
                for row_idx, point in enumerate(points, 2):
                    if use_polar:
                        data = [
                            point['point_index'],
                            point.get('r_coord'),
                            point.get('theta_coord'),
                            point.get('x_coord'),
                            point.get('y_coord'),
                            point.get('time_diff'),
                            point.get('stress_value'),
                            point.get('status'),
                            point.get('quality_score'),
                            point.get('snr'),
                            '是' if point.get('is_suspicious') else '',
                            point.get('measured_at'),
                            point.get('skip_reason', '')
                        ]
                    else:
                        data = [
                            point['point_index'],
                            point.get('x_coord'),
                            point.get('y_coord'),
                            point.get('time_diff'),
                            point.get('stress_value'),
                            point.get('status'),
                            point.get('quality_score'),
                            point.get('snr'),
                            '是' if point.get('is_suspicious') else '',
                            point.get('measured_at'),
                            point.get('skip_reason', '')
                        ]
                    
                    for col_idx, value in enumerate(data, 1):
                        ws_points.cell(row=row_idx, column=col_idx, value=value)
                
                # Sheet 3: 统计信息
                ws_stats = wb.create_sheet('统计信息')
                
                measured_points = [p for p in points if p.get('status') == 'measured']
                stress_values = [p['stress_value'] for p in measured_points if p.get('stress_value') is not None]
                
                stats_data = [
                    ('总测点数', len(points)),
                    ('已测量', len(measured_points)),
                    ('待测量', len([p for p in points if p.get('status') == 'pending'])),
                    ('已跳过', len([p for p in points if p.get('status') == 'skipped'])),
                    ('可疑点', len([p for p in points if p.get('is_suspicious')])),
                ]
                
                if stress_values:
                    stats_data.extend([
                        ('', ''),
                        ('应力统计', ''),
                        ('最小值(MPa)', min(stress_values)),
                        ('最大值(MPa)', max(stress_values)),
                        ('平均值(MPa)', np.mean(stress_values)),
                        ('标准差(MPa)', np.std(stress_values)),
                        ('范围(MPa)', max(stress_values) - min(stress_values)),
                    ])
                
                for row_idx, (key, value) in enumerate(stats_data, 1):
                    ws_stats.cell(row=row_idx, column=1, value=key)
                    if isinstance(value, float):
                        ws_stats.cell(row=row_idx, column=2, value=round(value, 2))
                    else:
                        ws_stats.cell(row=row_idx, column=2, value=value)
                
                # Sheet 4: 配置信息（形状和布点参数）
                ws_config = wb.create_sheet('配置信息')
                
                config_data = []
                
                # 形状配置
                config_data.append(('--- 形状配置 ---', ''))
                shape_type_map = {'rectangle': '矩形', 'circle': '圆形', 'polygon': '多边形'}
                shape_type = shape_config.get('type', '')
                config_data.append(('形状类型', shape_type_map.get(shape_type, shape_type)))
                
                if shape_type == 'rectangle':
                    config_data.append(('宽度(mm)', shape_config.get('width', '')))
                    config_data.append(('高度(mm)', shape_config.get('height', '')))
                elif shape_type == 'circle':
                    config_data.append(('半径(mm)', shape_config.get('radius', '')))
                    center = shape_config.get('center', {})
                    config_data.append(('圆心X(mm)', center.get('x', '')))
                    config_data.append(('圆心Y(mm)', center.get('y', '')))
                elif shape_type == 'polygon':
                    vertices = shape_config.get('vertices', [])
                    config_data.append(('顶点数', len(vertices)))
                    for i, v in enumerate(vertices):
                        config_data.append((f'顶点{i+1}', f"({v.get('x', 0)}, {v.get('y', 0)})"))
                
                # 孔洞信息
                modifiers = shape_config.get('modifiers', [])
                holes = [m for m in modifiers if m.get('op') == 'subtract']
                if holes:
                    config_data.append(('孔洞数量', len(holes)))
                    for i, hole in enumerate(holes, 1):
                        hole_shape = hole.get('shape', 'circle')
                        if hole_shape == 'circle':
                            config_data.append((f'孔洞{i}类型', '圆形'))
                            config_data.append((f'孔洞{i}圆心X(mm)', hole.get('centerX', '')))
                            config_data.append((f'孔洞{i}圆心Y(mm)', hole.get('centerY', '')))
                            config_data.append((f'孔洞{i}半径(mm)', hole.get('radius', '')))
                        elif hole_shape == 'rectangle':
                            config_data.append((f'孔洞{i}类型', '矩形'))
                            config_data.append((f'孔洞{i}中心X(mm)', hole.get('centerX', '')))
                            config_data.append((f'孔洞{i}中心Y(mm)', hole.get('centerY', '')))
                            config_data.append((f'孔洞{i}宽度(mm)', hole.get('width', '')))
                            config_data.append((f'孔洞{i}高度(mm)', hole.get('height', '')))
                
                config_data.append(('', ''))
                
                # 布点配置
                config_data.append(('--- 布点配置 ---', ''))
                layout_config = exp_data.get('config_snapshot', {}).get('layout', {})
                layout_type_map = {'grid': '网格布点', 'polar': '极坐标布点', 'custom': '自定义布点', 'adaptive': '自适应布点'}
                layout_type = layout_config.get('type', '')
                config_data.append(('布点方式', layout_type_map.get(layout_type, layout_type)))
                
                # 获取布点参数（可能在params子字典中）
                layout_params = layout_config.get('params', {})
                
                # 如果 params 为空，尝试从测点数据反推
                if not layout_params and points and layout_type == 'grid':
                    layout_params = self._infer_grid_params_from_points(points, shape_config)
                
                if layout_type == 'grid':
                    # 行数和列数（优先显示）
                    rows = layout_params.get('rows', '')
                    cols = layout_params.get('cols', '')
                    if rows:
                        config_data.append(('行数', rows))
                    if cols:
                        config_data.append(('列数', cols))
                    
                    # 边距可能在margins子字典中，也可能直接在params中
                    margins = layout_params.get('margins', {})
                    if margins:
                        config_data.append(('左边距(mm)', margins.get('left', '')))
                        config_data.append(('右边距(mm)', margins.get('right', '')))
                        config_data.append(('上边距(mm)', margins.get('top', '')))
                        config_data.append(('下边距(mm)', margins.get('bottom', '')))
                    else:
                        # 尝试直接从params读取
                        margin_left = layout_params.get('margin_left', '')
                        margin_right = layout_params.get('margin_right', '')
                        margin_top = layout_params.get('margin_top', '')
                        margin_bottom = layout_params.get('margin_bottom', '')
                        if margin_left or margin_right or margin_top or margin_bottom:
                            config_data.append(('左边距(mm)', margin_left))
                            config_data.append(('右边距(mm)', margin_right))
                            config_data.append(('上边距(mm)', margin_top))
                            config_data.append(('下边距(mm)', margin_bottom))
                        
                elif layout_type == 'polar':
                    config_data.append(('半径步长(mm)', layout_params.get('radius_step') or layout_params.get('r_step', '')))
                    config_data.append(('角度步长(°)', layout_params.get('angle_step', '')))
                    config_data.append(('起始半径(mm)', layout_params.get('start_radius') or layout_params.get('r_start', '')))
                    config_data.append(('圆心X(mm)', layout_params.get('center_x', '')))
                    config_data.append(('圆心Y(mm)', layout_params.get('center_y', '')))
                    config_data.append(('每层点数', layout_params.get('points_per_ring', '')))
                
                config_data.append(('', ''))
                config_data.append(('总测点数', len(points)))
                
                # 云图配置（如果有）
                metadata = result['data'].get('metadata')
                if metadata:
                    config_data.append(('', ''))
                    config_data.append(('--- 云图配置 ---', ''))
                    config_data.append(('插值方法', metadata.get('interpolation_method', '')))
                    config_data.append(('网格分辨率', metadata.get('grid_resolution', '')))
                    config_data.append(('色标范围最小值(MPa)', metadata.get('vmin', '')))
                    config_data.append(('色标范围最大值(MPa)', metadata.get('vmax', '')))
                
                for row_idx, (key, value) in enumerate(config_data, 1):
                    cell = ws_config.cell(row=row_idx, column=1, value=key)
                    if key.startswith('---'):
                        cell.font = Font(bold=True)
                    ws_config.cell(row=row_idx, column=2, value=value)
            
            # 保存
            wb.save(output_path)
            
            format_name = '单表' if single_sheet else '多表'
            return {
                "success": True,
                "error_code": 0,
                "file_path": output_path,
                "message": f"已导出Excel文件（{format_name}格式）"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error_code": 5011,
                "message": f"Excel导出失败: {str(e)}"
            }
    
    def export_to_hdf5(self, exp_id: str, output_path: str = None,
                      include_waveforms: bool = True) -> Dict[str, Any]:
        """
        导出为HDF5文件
        
        Args:
            exp_id: 实验ID
            output_path: 输出路径 (可选)
            include_waveforms: 是否包含波形数据
        
        Returns:
            dict: {"success": bool, "file_path": str}
        """
        try:
            import h5py
            
            # 加载实验数据
            result = self.db.load_experiment(exp_id)
            if not result['success']:
                return result
            
            exp_data = result['data']['experiment']
            points = result['data']['points']
            
            # 生成输出路径
            if output_path is None:
                output_dir = 'data/uniaxial_field/exports'
                os.makedirs(output_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(output_dir, f'{exp_id}_export_{timestamp}.h5')
            
            # 源HDF5文件
            source_hdf5 = FieldExperimentHDF5(exp_id)
            
            with h5py.File(output_path, 'w') as f:
                # 保存元数据
                meta_grp = f.create_group('metadata')
                for key, value in exp_data.items():
                    if value is not None and not isinstance(value, (dict, list)):
                        meta_grp.attrs[key] = str(value)
                
                # 保存配置
                config_grp = f.create_group('config')
                config_grp.attrs['shape_config'] = json.dumps(exp_data.get('shape_config', {}))
                config_grp.attrs['point_layout'] = json.dumps(exp_data.get('point_layout', []))
                
                # 保存测点数据
                points_grp = f.create_group('points')
                
                # 创建数据集
                n_points = len(points)
                points_grp.create_dataset('point_index', data=[p['point_index'] for p in points])
                points_grp.create_dataset('x_coord', data=[p.get('x_coord', 0) for p in points])
                points_grp.create_dataset('y_coord', data=[p.get('y_coord', 0) for p in points])
                points_grp.create_dataset('time_diff', data=[p.get('time_diff', 0) or 0 for p in points])
                points_grp.create_dataset('stress_value', data=[p.get('stress_value', 0) or 0 for p in points])
                points_grp.create_dataset('quality_score', data=[p.get('quality_score', 0) or 0 for p in points])
                points_grp.create_dataset('snr', data=[p.get('snr', 0) or 0 for p in points])
                
                # 保存波形数据
                if include_waveforms and source_hdf5.file_exists():
                    waveforms_grp = f.create_group('waveforms')
                    
                    # 复制基准波形
                    baseline_result = source_hdf5.load_baseline()
                    if baseline_result['success']:
                        baseline_grp = waveforms_grp.create_group('baseline')
                        wf = baseline_result['data']['waveform']
                        baseline_grp.create_dataset('time', data=wf['time'], compression='gzip')
                        baseline_grp.create_dataset('voltage', data=wf['voltage'], compression='gzip')
                    
                    # 复制测点波形
                    for point in points:
                        if point.get('status') == 'measured':
                            wf_result = source_hdf5.load_point_waveform(point['point_index'])
                            if wf_result['success']:
                                point_grp = waveforms_grp.create_group(f"point_{point['point_index']:03d}")
                                wf = wf_result['data']['waveform']
                                point_grp.create_dataset('time', data=wf['time'], compression='gzip')
                                point_grp.create_dataset('voltage', data=wf['voltage'], compression='gzip')
            
            return {
                "success": True,
                "error_code": 0,
                "file_path": output_path,
                "message": f"已导出HDF5文件"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error_code": 5021,
                "message": f"HDF5导出失败: {str(e)}"
            }
